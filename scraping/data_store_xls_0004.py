# 1) 지난 시간에 작성한 meltop100.csv 파일을 읽어,
#    meltop100.xlsx 로 저장하시오.
#    (단, 랭킹, 좋아요, 좋아요차이 컬럼은 숫자형식으로 저장 할 것!)
# 2) 멜론 Top100 곡들의 `앨범 재킷파일`을 다운받아,
#    meltop100.xlsx 파일의 두번째 시트에 랭킹순으로 작성하시오.
# 	(단, 이미지파일의 크기는 축소하여 보기 좋게 작성 할 것!)
# 3) 상위 Top10의 `좋아요 수`는 BarChart로, 
#    `좋아요 차이 수`는 ScatterChart로 세번째 시트에 작성하시오.

import csv
import codecs
import openpyxl
from openpyxl.chart import Reference, BarChart, Series, ScatterChart

## 멜론 자켓 이미지 얻어오는 함수.
def get_jacket():
    import requests
    from bs4 import BeautifulSoup

    url = "https://www.melon.com/chart/index.htm"

    headers = {"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_12_0) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/71.0.3578.98 Safari/537.36"}

    res = requests.get(url, headers=headers)
    soup = BeautifulSoup(res.text, 'html.parser')

    selector = ".wrap a img"
    images = soup.select(selector)

    for i, image in enumerate(images):
        image_url = image.get('src')
        img = requests.get(image_url).content

        saveFile = "./results/images/"+str(i+1)+".jpg"
        with open(saveFile, mode="wb") as file:
            file.write(img)
            print("----------"+saveFile+"----------")

## codecs 모듈로 csv파일 읽어 들이기
csvFile = "./results/data/melon_top_100_list.csv"
csvFile_encoded = codecs.open(csvFile, "r", "utf-8")
reader = csv.reader(csvFile_encoded, delimiter=',', quotechar='"')

## 1번
book = openpyxl.Workbook()
sheet1 = book.active

p = list(reader)
pp = len(p)

for i, cells in enumerate(p):
    for j, cell in enumerate(cells):
        print(i, " ---- ", j, " === ", cell)
        if i == 0:
            sheet1.cell(row = (i + 1), column = (j + 1)).value = cell
        elif i == pp - 1:
            if j > 2:
                sheet1.cell(row = (i + 1), column = (j + 1)).number_format
                sheet1.cell(row = (i + 1), column = (j + 1)).value = int(cell)
            else:
                sheet1.cell(row = (i + 1), column = (j + 1)).value = cell
        else:
            if j == 0 or j > 2:
                sheet1.cell(row = (i + 1), column = (j + 1)).number_format
                sheet1.cell(row = (i + 1), column = (j + 1)).value = int(cell)
            else:
                sheet1.cell(row = (i + 1), column = (j + 1)).value = cell
        
book.save("./results/data/meltop100.xlsx")
print("Excel file has been saved!")

## 2번
# get_jacket()  # 엑셀에 이미지 넣기 전에 먼저 실행하여 이미지를 저장.
book = openpyxl.load_workbook("./results/data/meltop100.xlsx")
sheet2 = book.create_sheet()

rows = ["A","C","E","G","I","K", "M", "O", "Q", "S"]

j = 0
for i in range(100):
    k = (i // 10)
    if j == 10: j = 0
    imgFile = './results/images/'+str(i+1)+'.jpg'
    img = openpyxl.drawing.image.Image(imgFile)
    anchor = rows[k]+str((j * 7) + 1)
    sheet2.add_image(img,anchor)
    print("----- ", anchor, " image -----")
    j += 1

book.save("./results/data/meltop100.xlsx")
print("Excel file has been saved!")

## 3번.
book = openpyxl.load_workbook("./results/data/meltop100.xlsx")
sheet3 = book.create_sheet()

data = []
for r in sheet1.rows:
    data.append([ r[1].value, r[3].value, r[4].value ])

### 3-1. 막대그래프 - 좋아요
datax = Reference(sheet1, min_col=4, 
		min_row=2, max_col=4, max_row=11)
categs = Reference(sheet1, min_col=1,
				 min_row=2, max_row=11)

chart = BarChart()
chart.title = '좋아요 수'
chart.add_data(data=datax)
chart.set_categories(categs)
chart.legend = None
chart.varyColors = True

sheet3.add_chart(chart, "B5")

book.save("./results/data/meltop100.xlsx")
print("Excel file has been saved!")

### 3-2. 산포도 - 좋아요 차이 수
book = openpyxl.load_workbook("./results/data/meltop100.xlsx")
sheet1 = book.worksheets[0]
sheet3 = book.worksheets[2]

chart1 = ScatterChart()
chart1.style = 13
chart1.x_axis.title = 'Size'
chart1.y_axis.title = 'Percentage'
chart1.title = '좋아요 차이 수'

xvalues = Reference(sheet1, min_col=1, min_row=2, max_row=11)
values = Reference(sheet1,min_col=5, min_row=2, max_row=11)
series = Series(values, xvalues, title_from_data=True)
chart1.series.append(series)

sheet3.add_chart(chart1, "A23")

book.save("./results/data/meltop100.xlsx")
print("Excel file has been saved!")

